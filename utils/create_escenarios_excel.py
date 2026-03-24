"""Genera el Excel de escenarios de impuestos con los 10 casos definidos.

Uso:
    cd cfdi-automation
    python utils/create_escenarios_excel.py

Crea (o sobreescribe):
    data/conceptos-escenario-ppd-1.xlsx

La estructura generada es compatible con la hoja 'conceptos' y
'impuestos-conceptos' existentes — solo agrega la columna 'Id Escenario'
al inicio de cada hoja.

ESCENARIOS GENERADOS
────────────────────
Esc 1  → IVA 16
Esc 2  → IVA 0
Esc 3  → IVA 16 + IVA 0
Esc 4  → IEPS 0.53 + IEPS 0.30 + IEPS 0.265
Esc 5  → IVA 16 + ISR Retención 0.15
Esc 6  → IVA 16 + Retención IEPS 0.53 + IEPS 0.30
Esc 7  → IVA 16 + IVA 0 + IEPS 0.08
Esc 8  → IVA 16 + IVA 0 + IEPS 0.08 + IEPS 0.09
Esc 9  → IVA 16 + IVA 16 Ret. + ISR 0.15 Ret.
Esc 10 → IVA 16 + IVA 0 + IEPS 0.09 + IVA 16 Ret. + IEPS 0.09 Ret. + ISR 16 Ret.
"""
import os
import sys

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ── Ruta de salida ─────────────────────────────────────────────────────────────
_BASE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
_OUT = os.path.join(_BASE, "data", "conceptos-escenario-ppd-1.xlsx")

# ── Estilos ────────────────────────────────────────────────────────────────────
_HDR_FILL   = PatternFill("solid", fgColor="2D5F9E")
_HDR_FONT   = Font(bold=True, color="FFFFFF", size=10)
_ESC_FILL   = {
    i: PatternFill("solid", fgColor=c)
    for i, c in enumerate([
        # 1-10 (existentes)
        "D9EAF7","D9F7E8","FFF9D9","F7E8D9","EFD9F7",
        "D9F7F7","F7D9EF","E8F7D9","F7F0D9","F0D9F7",
        # 11-18 (nuevos — Hyatt)
        "E8D9F7","D9F7E0","F7EFD9","D9EFF7",
        "F7D9D9","D9F7D9","F7F7D9","D9E8F7",
    ], start=1)
}
_THIN = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
_ALIGN = Alignment(horizontal="left", vertical="center")
_ALIGN_C = Alignment(horizontal="center", vertical="center")


def _hdr(ws, row, col, value):
    c = ws.cell(row=row, column=col, value=value)
    c.font = _HDR_FONT
    c.fill = _HDR_FILL
    c.alignment = _ALIGN_C
    c.border = _THIN


def _cell(ws, row, col, value, fill=None):
    c = ws.cell(row=row, column=col, value=value)
    c.alignment = _ALIGN
    c.border = _THIN
    if fill:
        c.fill = fill


# ═══════════════════════════════════════════════════════════════════════════════
# DATOS DE LOS EMISORES
# ═══════════════════════════════════════════════════════════════════════════════

# Emisor por defecto (escenarios 1-10): celdas vacías → el test usa config.ini
_EMISOR_DEFAULT = {"rfc": "", "sucursal": "", "cc": ""}

# Emisor Hyatt (escenarios 11-18)
_EMISOR_HYATT = {
    "rfc":      "TPA110608SW9",
    "sucursal": "Hyatt Place Tijuana",
    "cc":       "TEST OPERA CLOUD",
}


# ═══════════════════════════════════════════════════════════════════════════════
# DATOS DE LOS ESCENARIOS
# ═══════════════════════════════════════════════════════════════════════════════

_ESCENARIOS = [
    # ─── Grupo 1: Emisor NID200929V26 (usa config.ini, celdas vacías en Excel) ───
    {"id": 1,  "nombre": "IVA 16",                                  **_EMISOR_DEFAULT},
    {"id": 2,  "nombre": "IVA 0",                                   **_EMISOR_DEFAULT},
    {"id": 3,  "nombre": "IVA 16 e IVA 0",                         **_EMISOR_DEFAULT},
    {"id": 4,  "nombre": "IEPS 0.53 IEPS 0.30 IEPS 0.265",         **_EMISOR_DEFAULT},
    {"id": 5,  "nombre": "IVA 16 e ISR Retencion 0.15",            **_EMISOR_DEFAULT},
    {"id": 6,  "nombre": "IVA 16 Retencion IEPS 53 IEPS 30",       **_EMISOR_DEFAULT},
    {"id": 7,  "nombre": "IVA 16 IVA 0 IEPS 0.08",                 **_EMISOR_DEFAULT},
    {"id": 8,  "nombre": "IVA 16 IVA 0 IEPS 0.08 IEPS 0.09",      **_EMISOR_DEFAULT},
    {"id": 9,  "nombre": "IVA 16 IVA 16 Ret ISR 0.15 Ret",         **_EMISOR_DEFAULT},
    {"id": 10, "nombre": "IVA 16 IVA 0 IEPS 0.09 Ret IVA Ret IEPS Ret ISR", **_EMISOR_DEFAULT},
    # ─── Grupo 2: Emisor TPA110608SW9 (Hyatt Place Tijuana) ───
    {"id": 11, "nombre": "IVA 8",                                   **_EMISOR_HYATT},
    {"id": 12, "nombre": "IVA 8 e IVA 0",                          **_EMISOR_HYATT},
    {"id": 13, "nombre": "IVA 16 IVA 8 e IVA 0",                   **_EMISOR_HYATT},
    {"id": 14, "nombre": "IVA 8 IVA 0 IEPS 0.08 Ret IVA Ret IEPS ISH 8", **_EMISOR_HYATT},
    {"id": 15, "nombre": "IVA 16 IEPS 0.09 IEPS 0.30 Ret IVA ISH 0.09 ISH 0.08", **_EMISOR_HYATT},
    {"id": 16, "nombre": "IVA 16 IVA 8 IVA 0 IEPS 0.08 IEPS 0.53 IEPS 30", **_EMISOR_HYATT},
    {"id": 17, "nombre": "IVA 8 IVA 0 IEPS 0.08 IEPS 0.53 Ret IEPS ISH 0.08", **_EMISOR_HYATT},
    {"id": 18, "nombre": "IVA 16 IVA 8 IVA 0 IEPS 0.53 IEPS 0.265 Ret IEPS Ret ISR Ret ISH", **_EMISOR_HYATT},
]

# Concepto base: el mismo para todos los escenarios
# (Id Concepto siempre = 1 dentro de cada escenario)
_CONCEPTO_BASE = {
    "Cantidad":                 1,
    "Clave Unidad":             "H87",
    "Descripción":              "Servicio de prueba",
    "Clave Producto/Servicio":  "84111506",
    "Valor Unitario":           100.00,
    "Objeto Impuesto":          "02",
    "Id Concepto":              1,
}


def _I(esc_id, conc_id, tipo, lf, ret_tras, factor, tasa, nombre=""):
    """Helper para construir una fila de impuesto."""
    return {
        "Id Escenario":       esc_id,
        "Id Concepto":        conc_id,
        "Tipo Impuesto":      tipo,
        "Local o Federal":    lf,
        "Retencion o Traslado": ret_tras,
        "Tipo Factor":        factor,
        "Tasa o Cuota":       tasa,
        "Nombre Impuesto":    nombre or f"{tipo} {tasa}",
    }


# Impuestos por escenario  (I = traslado, R = retención)
_IMPUESTOS = [
    # Esc 1: IVA 16
    _I(1,  1, "IVA",  "FEDERAL", "TRASLADO",  "TASA", 0.160, "IVA 16%"),

    # Esc 2: IVA 0
    _I(2,  1, "IVA",  "FEDERAL", "TRASLADO",  "TASA", 0.000, "IVA 0%"),

    # Esc 3: IVA 16 + IVA 0
    _I(3,  1, "IVA",  "FEDERAL", "TRASLADO",  "TASA", 0.160, "IVA 16%"),
    _I(3,  1, "IVA",  "FEDERAL", "TRASLADO",  "TASA", 0.000, "IVA 0%"),

    # Esc 4: IEPS 0.53 + IEPS 0.30 + IEPS 0.265
    _I(4,  1, "IEPS", "FEDERAL", "TRASLADO",  "TASA", 0.530, "IEPS 53%"),
    _I(4,  1, "IEPS", "FEDERAL", "TRASLADO",  "TASA", 0.300, "IEPS 30%"),
    _I(4,  1, "IEPS", "FEDERAL", "TRASLADO",  "TASA", 0.265, "IEPS 26.5%"),

    # Esc 5: IVA 16 + ISR Retención 0.15
    _I(5,  1, "IVA",  "FEDERAL", "TRASLADO",  "TASA", 0.160, "IVA 16%"),
    _I(5,  1, "ISR",  "FEDERAL", "RETENCION", "TASA", 0.150, "ISR Ret 15%"),

    # Esc 6: IVA 16 + Retención IEPS 0.53 + IEPS 0.30
    _I(6,  1, "IVA",  "FEDERAL", "TRASLADO",  "TASA", 0.160, "IVA 16%"),
    _I(6,  1, "IEPS", "FEDERAL", "RETENCION", "TASA", 0.530, "Ret IEPS 53%"),
    _I(6,  1, "IEPS", "FEDERAL", "TRASLADO",  "TASA", 0.300, "IEPS 30%"),

    # Esc 7: IVA 16 + IVA 0 + IEPS 0.08
    _I(7,  1, "IVA",  "FEDERAL", "TRASLADO",  "TASA", 0.160, "IVA 16%"),
    _I(7,  1, "IVA",  "FEDERAL", "TRASLADO",  "TASA", 0.000, "IVA 0%"),
    _I(7,  1, "IEPS", "FEDERAL", "TRASLADO",  "TASA", 0.080, "IEPS 8%"),

    # Esc 8: IVA 16 + IVA 0 + IEPS 0.08 + IEPS 0.09
    _I(8,  1, "IVA",  "FEDERAL", "TRASLADO",  "TASA", 0.160, "IVA 16%"),
    _I(8,  1, "IVA",  "FEDERAL", "TRASLADO",  "TASA", 0.000, "IVA 0%"),
    _I(8,  1, "IEPS", "FEDERAL", "TRASLADO",  "TASA", 0.080, "IEPS 8%"),
    _I(8,  1, "IEPS", "FEDERAL", "TRASLADO",  "TASA", 0.090, "IEPS 9%"),

    # Esc 9: IVA 16 + IVA 16 Ret + ISR 0.15 Ret
    _I(9,  1, "IVA",  "FEDERAL", "TRASLADO",  "TASA", 0.160, "IVA 16%"),
    _I(9,  1, "IVA",  "FEDERAL", "RETENCION", "TASA", 0.160, "Ret IVA 16%"),
    _I(9,  1, "ISR",  "FEDERAL", "RETENCION", "TASA", 0.150, "Ret ISR 15%"),

    # Esc 10: IVA 16 + IVA 0 + IEPS 0.09 + Ret IVA 16 + Ret IEPS 0.09 + Ret ISR 16
    _I(10, 1, "IVA",  "FEDERAL", "TRASLADO",  "TASA", 0.160, "IVA 16%"),
    _I(10, 1, "IVA",  "FEDERAL", "TRASLADO",  "TASA", 0.000, "IVA 0%"),
    _I(10, 1, "IEPS", "FEDERAL", "TRASLADO",  "TASA", 0.090, "IEPS 9%"),
    _I(10, 1, "IVA",  "FEDERAL", "RETENCION", "TASA", 0.160, "Ret IVA 16%"),
    _I(10, 1, "IEPS", "FEDERAL", "RETENCION", "TASA", 0.090, "Ret IEPS 9%"),
    _I(10, 1, "ISR",  "FEDERAL", "RETENCION", "TASA", 0.160, "Ret ISR 16%"),

    # ─── Nuevos escenarios Hyatt Place Tijuana (RFC TPA110608SW9) ───

    # Esc 11: IVA 8
    _I(11, 1, "IVA",  "FEDERAL", "TRASLADO",  "TASA", 0.080, "IVA 8%"),

    # Esc 12: IVA 8 + IVA 0
    _I(12, 1, "IVA",  "FEDERAL", "TRASLADO",  "TASA", 0.080, "IVA 8%"),
    _I(12, 1, "IVA",  "FEDERAL", "TRASLADO",  "TASA", 0.000, "IVA 0%"),

    # Esc 13: IVA 16 + IVA 8 + IVA 0
    _I(13, 1, "IVA",  "FEDERAL", "TRASLADO",  "TASA", 0.160, "IVA 16%"),
    _I(13, 1, "IVA",  "FEDERAL", "TRASLADO",  "TASA", 0.080, "IVA 8%"),
    _I(13, 1, "IVA",  "FEDERAL", "TRASLADO",  "TASA", 0.000, "IVA 0%"),

    # Esc 14: IVA 8 + IVA 0 + IEPS 0.08 (traslado) + Ret IVA 0.08 + Ret IEPS 0.08 + ISH 0.08 (local traslado)
    _I(14, 1, "IVA",  "FEDERAL", "TRASLADO",  "TASA", 0.080, "IVA 8%"),
    _I(14, 1, "IVA",  "FEDERAL", "TRASLADO",  "TASA", 0.000, "IVA 0%"),
    _I(14, 1, "IEPS", "FEDERAL", "TRASLADO",  "TASA", 0.080, "IEPS 8%"),
    _I(14, 1, "IVA",  "FEDERAL", "RETENCION", "TASA", 0.080, "Ret IVA 8%"),
    _I(14, 1, "IEPS", "FEDERAL", "RETENCION", "TASA", 0.080, "Ret IEPS 8%"),
    _I(14, 1, "ISH",  "LOCAL",   "TRASLADO",  "TASA", 0.080, "ISH 8%"),

    # Esc 15: IVA 16 + IEPS 0.09 + IEPS 0.30 + Ret IVA 0.08 + ISH 0.09 + ISH 0.08
    _I(15, 1, "IVA",  "FEDERAL", "TRASLADO",  "TASA", 0.160, "IVA 16%"),
    _I(15, 1, "IEPS", "FEDERAL", "TRASLADO",  "TASA", 0.090, "IEPS 9%"),
    _I(15, 1, "IEPS", "FEDERAL", "TRASLADO",  "TASA", 0.300, "IEPS 30%"),
    _I(15, 1, "IVA",  "FEDERAL", "RETENCION", "TASA", 0.080, "Ret IVA 8%"),
    _I(15, 1, "ISH",  "LOCAL",   "TRASLADO",  "TASA", 0.090, "ISH 9%"),
    _I(15, 1, "ISH",  "LOCAL",   "TRASLADO",  "TASA", 0.080, "ISH 8%"),

    # Esc 16: IVA 16 + IVA 8 + IVA 0 + IEPS 0.08 + IEPS 0.53 + IEPS 0.30 (todos traslados)
    _I(16, 1, "IVA",  "FEDERAL", "TRASLADO",  "TASA", 0.160, "IVA 16%"),
    _I(16, 1, "IVA",  "FEDERAL", "TRASLADO",  "TASA", 0.080, "IVA 8%"),
    _I(16, 1, "IVA",  "FEDERAL", "TRASLADO",  "TASA", 0.000, "IVA 0%"),
    _I(16, 1, "IEPS", "FEDERAL", "TRASLADO",  "TASA", 0.080, "IEPS 8%"),
    _I(16, 1, "IEPS", "FEDERAL", "TRASLADO",  "TASA", 0.530, "IEPS 53%"),
    _I(16, 1, "IEPS", "FEDERAL", "TRASLADO",  "TASA", 0.300, "IEPS 30%"),

    # Esc 17: IVA 8 + IVA 0 + IEPS 0.08 + IEPS 0.53 + Ret IEPS 0.08 + ISH 0.08
    _I(17, 1, "IVA",  "FEDERAL", "TRASLADO",  "TASA", 0.080, "IVA 8%"),
    _I(17, 1, "IVA",  "FEDERAL", "TRASLADO",  "TASA", 0.000, "IVA 0%"),
    _I(17, 1, "IEPS", "FEDERAL", "TRASLADO",  "TASA", 0.080, "IEPS 8%"),
    _I(17, 1, "IEPS", "FEDERAL", "TRASLADO",  "TASA", 0.530, "IEPS 53%"),
    _I(17, 1, "IEPS", "FEDERAL", "RETENCION", "TASA", 0.080, "Ret IEPS 8%"),
    _I(17, 1, "ISH",  "LOCAL",   "TRASLADO",  "TASA", 0.080, "ISH 8%"),

    # Esc 18: IVA 16 + IVA 8 + IVA 0 + IEPS 0.08 + IEPS 0.53 + IEPS 0.265 + Ret IEPS 0.08 + Ret ISR 0.015 + Ret ISH 0.05
    _I(18, 1, "IVA",  "FEDERAL", "TRASLADO",  "TASA", 0.160, "IVA 16%"),
    _I(18, 1, "IVA",  "FEDERAL", "TRASLADO",  "TASA", 0.080, "IVA 8%"),
    _I(18, 1, "IVA",  "FEDERAL", "TRASLADO",  "TASA", 0.000, "IVA 0%"),
    _I(18, 1, "IEPS", "FEDERAL", "TRASLADO",  "TASA", 0.080, "IEPS 8%"),
    _I(18, 1, "IEPS", "FEDERAL", "TRASLADO",  "TASA", 0.530, "IEPS 53%"),
    _I(18, 1, "IEPS", "FEDERAL", "TRASLADO",  "TASA", 0.265, "IEPS 26.5%"),
    _I(18, 1, "IEPS", "FEDERAL", "RETENCION", "TASA", 0.080, "Ret IEPS 8%"),
    _I(18, 1, "ISR",  "FEDERAL", "RETENCION", "TASA", 0.015, "Ret ISR 1.5%"),
    _I(18, 1, "ISH",  "LOCAL",   "RETENCION", "TASA", 0.050, "Ret ISH 5%"),
]


# ═══════════════════════════════════════════════════════════════════════════════
# CONSTRUCCIÓN DEL EXCEL
# ═══════════════════════════════════════════════════════════════════════════════

def build_excel(path: str, preserve_pagos: bool = True):
    """Construye el Excel completo con las 4 hojas requeridas.

    Si el archivo ya existe y preserve_pagos=True, intenta preservar la hoja
    'pagos' sin modificarla.
    """
    os.makedirs(os.path.dirname(path), exist_ok=True)

    # Leer datos de pagos del archivo existente si aplica
    pagos_rows = []
    pagos_headers = []
    if preserve_pagos and os.path.exists(path):
        try:
            wb_old = openpyxl.load_workbook(path, data_only=True)
            if "pagos" in wb_old.sheetnames:
                ws_p = wb_old["pagos"]
                for row in ws_p.iter_rows(values_only=True):
                    if pagos_headers:
                        pagos_rows.append(list(row))
                    else:
                        pagos_headers = [str(v) if v is not None else "" for v in row]
            wb_old.close()
        except Exception:
            pass

    wb = openpyxl.Workbook()

    # ── Hoja 1: escenarios ────────────────────────────────────────────────────
    ws_e = wb.active
    ws_e.title = "escenarios"
    esc_headers = ["Id Escenario", "Nombre Escenario", "RFC Emisor", "Sucursal", "CC"]
    esc_widths   = [14, 52, 18, 26, 22]
    for col, (h, w) in enumerate(zip(esc_headers, esc_widths), start=1):
        _hdr(ws_e, 1, col, h)
        ws_e.column_dimensions[openpyxl.utils.get_column_letter(col)].width = w
    for esc in _ESCENARIOS:
        fill = _ESC_FILL.get(esc["id"])
        r = esc["id"] + 1
        _cell(ws_e, r, 1, esc["id"],         fill)
        _cell(ws_e, r, 2, esc["nombre"],      fill)
        _cell(ws_e, r, 3, esc.get("rfc",  ""), fill)
        _cell(ws_e, r, 4, esc.get("sucursal", ""), fill)
        _cell(ws_e, r, 5, esc.get("cc",   ""), fill)

    # ── Hoja 2: conceptos ─────────────────────────────────────────────────────
    ws_c = wb.create_sheet("conceptos")
    conc_headers = [
        "Id Escenario", "Id Concepto", "Cantidad",
        "Clave Unidad", "Descripción", "Clave Producto/Servicio",
        "Valor Unitario", "Objeto Impuesto",
    ]
    widths_c = [14, 12, 10, 14, 28, 26, 14, 22]
    for col, (h, w) in enumerate(zip(conc_headers, widths_c), start=1):
        _hdr(ws_c, 1, col, h)
        ws_c.column_dimensions[openpyxl.utils.get_column_letter(col)].width = w

    for esc in _ESCENARIOS:
        r = esc["id"] + 1
        fill = _ESC_FILL.get(esc["id"])
        vals = [
            esc["id"],
            _CONCEPTO_BASE["Id Concepto"],
            _CONCEPTO_BASE["Cantidad"],
            _CONCEPTO_BASE["Clave Unidad"],
            _CONCEPTO_BASE["Descripción"],
            _CONCEPTO_BASE["Clave Producto/Servicio"],
            _CONCEPTO_BASE["Valor Unitario"],
            _CONCEPTO_BASE["Objeto Impuesto"],
        ]
        for col, v in enumerate(vals, start=1):
            _cell(ws_c, r, col, v, fill)

    # ── Hoja 3: impuestos-conceptos ───────────────────────────────────────────
    ws_i = wb.create_sheet("impuestos-conceptos")
    imp_headers = [
        "Id Escenario", "Id Concepto", "Tipo Impuesto",
        "Local o Federal", "Retencion o Traslado",
        "Tipo Factor", "Tasa o Cuota", "Nombre Impuesto",
    ]
    widths_i = [14, 12, 14, 16, 22, 12, 14, 22]
    for col, (h, w) in enumerate(zip(imp_headers, widths_i), start=1):
        _hdr(ws_i, 1, col, h)
        ws_i.column_dimensions[openpyxl.utils.get_column_letter(col)].width = w

    for row_idx, imp in enumerate(_IMPUESTOS, start=2):
        fill = _ESC_FILL.get(imp["Id Escenario"])
        vals = [
            imp["Id Escenario"],
            imp["Id Concepto"],
            imp["Tipo Impuesto"],
            imp["Local o Federal"],
            imp["Retencion o Traslado"],
            imp["Tipo Factor"],
            imp["Tasa o Cuota"],
            imp["Nombre Impuesto"],
        ]
        for col, v in enumerate(vals, start=1):
            _cell(ws_i, row_idx, col, v, fill)

    # ── Hoja 4: pagos ─────────────────────────────────────────────────────────
    # Siempre se regenera con una fila por escenario.
    # Si el archivo ya existía, se preservan los valores reales de cada fila;
    # los escenarios nuevos reciben valores por defecto (0.0 para CP1/CP2).
    ws_p = wb.create_sheet("pagos")
    pagos_new_headers = ["ESCENARIO", "FECHA DE PAGO", "FORMA DE PAGO", "MONEDA DE PAGO", "CP1", "CP2"]
    pagos_new_widths  = [14, 18, 16, 18, 12, 12]
    for col, (h, w) in enumerate(zip(pagos_new_headers, pagos_new_widths), start=1):
        _hdr(ws_p, 1, col, h)
        ws_p.column_dimensions[openpyxl.utils.get_column_letter(col)].width = w

    # Construir mapa escenario → fila existente (si se preservaron datos)
    existing_pagos: dict = {}
    if pagos_headers and pagos_rows:
        hdrs_upper = [h.upper() for h in pagos_headers]
        esc_col_idx = next(
            (i for i, h in enumerate(hdrs_upper) if "ESCENARIO" in h or h == "ESC"),
            None,
        )
        if esc_col_idx is not None:
            for row_data in pagos_rows:
                try:
                    eid = int(row_data[esc_col_idx] or 0)
                    if eid:
                        existing_pagos[eid] = row_data
                except (TypeError, ValueError):
                    pass

    for row_idx, esc in enumerate(_ESCENARIOS, start=2):
        eid = esc["id"]
        fill = _ESC_FILL.get(eid)
        if eid in existing_pagos and pagos_headers:
            # Preservar valores previos, remapeando por nombre de columna
            old_row = existing_pagos[eid]
            old_map = dict(zip([h.upper() for h in pagos_headers], old_row))
            fecha   = old_map.get("FECHA DE PAGO") or old_map.get("FECHA") or "20/03/2026"
            forma   = old_map.get("FORMA DE PAGO") or old_map.get("FORMA") or "01"
            moneda  = old_map.get("MONEDA DE PAGO") or old_map.get("MONEDA") or "MXN"
            cp1     = old_map.get("CP1") or 0.0
            cp2     = old_map.get("CP2") or 0.0
        else:
            fecha, forma, moneda, cp1, cp2 = "20/03/2026", "01", "MXN", 0.0, 0.0

        for col, v in enumerate([eid, fecha, forma, moneda, cp1, cp2], start=1):
            _cell(ws_p, row_idx, col, v, fill)

    wb.save(path)
    return path


# ═══════════════════════════════════════════════════════════════════════════════
# MAIN
# ═══════════════════════════════════════════════════════════════════════════════

if __name__ == "__main__":
    import argparse

    parser = argparse.ArgumentParser(
        description="Genera el Excel de escenarios de impuestos."
    )
    parser.add_argument(
        "--output", "-o",
        default=_OUT,
        help=f"Ruta destino del Excel (default: {_OUT})",
    )
    parser.add_argument(
        "--no-preserve-pagos",
        action="store_true",
        help="No preservar la hoja 'pagos' del archivo existente",
    )
    args = parser.parse_args()

    out = build_excel(args.output, preserve_pagos=not args.no_preserve_pagos)
    print(f"✅ Excel generado exitosamente: {out}")
    print(f"   Escenarios: {len(_ESCENARIOS)} (esc 1-10: NID200929V26 | esc 11-18: TPA110608SW9)")
    print(f"   Filas de impuestos: {len(_IMPUESTOS)}")
