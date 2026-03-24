"""Excel I/O for test data and results."""
import os
from datetime import datetime
from dataclasses import dataclass, fields
from typing import List

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ── Result row dataclass ──────────────────────────────────────────────────────

@dataclass
class ResultRow:
    caso_prueba: str = ""
    resultado_esperado: str = ""
    rfc_emisor: str = ""
    rfc_receptor: str = ""
    sucursal: str = ""
    centro_consumo: str = ""
    datos_generales: str = ""
    uuid_relacionado: str = ""
    url_descarga_pdf: str = ""
    uuid_timbrado: str = ""
    resultado_obtenido: str = ""


_HEADERS = [
    "Caso de Prueba",
    "Resultado Esperado",
    "RFC Emisor",
    "RFC Receptor",
    "Sucursal",
    "Centro de Consumo",
    "Datos Generales",
    "UUID CFDI Relacionado",
    "URL de Descarga (PDF)",
    "UUID Timbrado",
    "Resultado Obtenido",
]

_HEADER_FILL = PatternFill("solid", fgColor="2D5F9E")
_PASS_FILL = PatternFill("solid", fgColor="C6EFCE")
_FAIL_FILL = PatternFill("solid", fgColor="FFC7CE")
_THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)


# ── Concepto / Impuesto dataclasses ──────────────────────────────────────────

@dataclass
class ConceptoRow:
    cantidad: float
    clave_unidad: str
    descripcion: str
    clave_producto: str
    valor_unitario: float
    objeto_impuesto: str
    id_concepto: int


@dataclass
class ImpuestoRow:
    id_concepto: int
    retencion_traslado: str
    clave_impuesto: str
    nombre_impuesto: str
    local_federal: str
    tipo_factor: str
    tasa_cuota: float


@dataclass
class PagoRow:
    """Datos del Complemento de Pago leídos desde la hoja 'pagos' del Excel."""
    id_escenario: int = 0      # columna ESCENARIO — clave de búsqueda
    fecha_pago: str = ""       # "2026-03-20" / "20/03/2026" / "20-03-2026"
    forma_pago: str = "01"     # código SAT, ej. "01", "03", "04"
    moneda_pago: str = "MXN"   # clave de moneda SAT
    cp1: float = 0.0           # importe del primer complemento de pago
    cp2: float = 0.0           # importe del segundo complemento de pago


@dataclass
class EscenarioData:
    """Agrupación de un escenario de prueba: conceptos + impuestos asociados.

    Cada instancia representa un caso de prueba independiente leído del Excel.
    El Excel es la única fuente de verdad; Python solo lee y ejecuta.

    Los campos rfc_emisor / sucursal / cc son opcionales: si están vacíos el
    test usa los valores del [emisor] de config.ini (comportamiento anterior).
    """
    id_escenario: int
    nombre: str
    conceptos: List[ConceptoRow]
    impuestos: List[ImpuestoRow]
    rfc_emisor: str = ""   # vacío → usar config [emisor] rfc
    sucursal: str = ""     # vacío → usar config [emisor] sucursal
    cc: str = ""           # vacío → usar config [emisor] cc

    def __str__(self) -> str:
        return f"Esc{self.id_escenario}_{self.nombre}"

    def __repr__(self) -> str:
        return (
            f"EscenarioData(id={self.id_escenario}, nombre={self.nombre!r}, "
            f"conceptos={len(self.conceptos)}, impuestos={len(self.impuestos)})"
        )


# ── Normalizers ───────────────────────────────────────────────────────────────

def _clean_str(value, default: str = "") -> str:
    if value is None:
        return default
    return str(value).strip()


def _normalize_local_federal(value) -> str:
    v = _clean_str(value).upper()
    if "LOCAL" in v:
        return "Local"
    if "FEDERAL" in v:
        return "Federal"
    return _clean_str(value).title()


def _normalize_retencion_traslado(value) -> str:
    v = _clean_str(value).upper()
    if "TRAS" in v:
        return "Traslado"
    if "RETEN" in v:
        return "Retención"
    return _clean_str(value).title()


def _normalize_tipo_factor(value) -> str:
    v = _clean_str(value).upper()
    if "TASA" in v:
        return "Tasa"
    if "CUOTA" in v:
        return "Cuota"
    if "EXENTO" in v:
        return "Exento"
    return _clean_str(value).title()


def _normalize_clave_impuesto(value) -> str:
    v = _clean_str(value).upper()
    if "IVA" in v:
        return "IVA"
    if "ISR" in v:
        return "ISR"
    if "IEPS" in v:
        return "IEPS"
    return _clean_str(value).upper()


def _normalize_objeto_impuesto(value) -> str:
    v = _clean_str(value).upper()
    if v.startswith("01"):
        return "01 No objeto de impuesto"
    if v.startswith("02"):
        return "02 Sí objeto de impuesto"
    if v.startswith("03"):
        return "03 Sí objeto del impuesto y no obligado al desglose"
    if v.startswith("04"):
        return "04 Sí objeto del impuesto y no causa impuesto"
    return _clean_str(value)

def _clean(value):
    return str(value).strip() if value else ""

def _norm_retencion(v):
    v = _clean(v).upper()
    if "TRAS" in v:
        return "Traslado"
    if "RET" in v:
        return "Retención"
    raise ValueError(f"Valor inválido en Retencion o Traslado: '{v}'")

def _norm_local(v):
    v = _clean(v).upper()
    if v == "LOCAL":
        return "Local"
    if v == "FEDERAL":
        return "Federal"
    raise ValueError(f"Valor inválido en Local o Federal: '{v}'")

def _norm_factor(v):
    v = _clean(v).upper()
    if v == "TASA":
        return "Tasa"
    if v == "CUOTA":
        return "Cuota"
    if v == "EXENTO":
        return "Exento"
    raise ValueError(f"Valor inválido en Tipo Factor: '{v}'")

def _get_first_present(d: dict, *keys, default=""):
    for key in keys:
        if key in d and d[key] not in (None, ""):
            return d[key]
    return default



# ── Reader ────────────────────────────────────────────────────────────────────

def read_conceptos(path: str):
    """Read conceptos and impuestos sheets from the given Excel file."""
    wb = openpyxl.load_workbook(path, data_only=True)

    # ---- conceptos sheet ----
    ws_c = wb["conceptos"]
    headers_c = [str(c.value).strip() if c.value is not None else "" for c in next(ws_c.iter_rows(min_row=1, max_row=1))]
    conceptos: List[ConceptoRow] = []

    for row in ws_c.iter_rows(min_row=2, values_only=True):
        if not any(row):
            continue

        r = dict(zip(headers_c, row))

        conceptos.append(ConceptoRow(
            cantidad=float(r.get("Cantidad", 1) or 1),
            clave_unidad=_clean_str(r.get("Clave Unidad", "")).upper(),
            descripcion=_clean_str(r.get("Descripción", "")).strip(),
            clave_producto=_clean_str(r.get("Clave Producto/Servicio", "")).strip(),
            valor_unitario=float(r.get("Valor Unitario", 0) or 0),
            objeto_impuesto=_normalize_objeto_impuesto(r.get("Objeto Impuesto", "02")),
            id_concepto=int(r.get("Id Concepto", 0) or 0),
        ))

    # ---- impuestos-conceptos sheet ----
    # ---- impuestos-conceptos sheet ----
    ws_i = wb["impuestos-conceptos"]
    headers_i = [str(c.value).strip() if c.value is not None else "" for c in next(ws_i.iter_rows(min_row=1, max_row=1))]
    impuestos: List[ImpuestoRow] = []

    # When the sheet has 'Id Escenario' column (multi-scenario Excel),
    # read_conceptos only returns impuestos for escenario 1 (backward compat).
    _has_esc_col = "Id Escenario" in headers_i

    for row in ws_i.iter_rows(min_row=2, values_only=True):
        if not any(row):
            continue

        r = dict(zip(headers_i, row))

        # Skip rows that belong to other scenarios
        if _has_esc_col:
            esc_id = r.get("Id Escenario")
            if esc_id is not None and int(esc_id) != 1:
                continue

        raw_retencion = _get_first_present(
            r,
            "Retencion o Traslado",
            "Retención o Traslado",
        )

        raw_tipo_impuesto = _get_first_present(r, "Tipo Impuesto")
        raw_local_federal = _get_first_present(r, "Local o Federal")
        raw_tipo_factor = _get_first_present(r, "Tipo Factor")
        raw_tasa = _get_first_present(r, "Tasa o Cuota", default=0)

        if not raw_retencion:
            raise ValueError(
                f"Excel inválido: columna 'Retencion o Traslado' vacía o no encontrada. "
                f"Headers detectados={headers_i} | fila={r}"
            )

        impuestos.append(ImpuestoRow(
            id_concepto=int(_get_first_present(r, "Id Concepto", default=0)),
            retencion_traslado=_norm_retencion(raw_retencion),
            clave_impuesto=_clean(raw_tipo_impuesto).upper(),
            nombre_impuesto=_clean(_get_first_present(r, "Nombre Impuesto")),
            local_federal=_norm_local(raw_local_federal),
            tipo_factor=_norm_factor(raw_tipo_factor),
            tasa_cuota=float(raw_tasa or 0),
        ))

    wb.close()
    return conceptos, impuestos


def read_escenarios(path: str) -> "List[EscenarioData]":
    """Lee todos los escenarios de prueba desde el Excel.

    Estructura esperada en el Excel:
    ─────────────────────────────────────────────────────────────────────────
    Hoja  'escenarios'         (opcional, para nombres de escenario)
        Id Escenario | Nombre Escenario

    Hoja  'conceptos'          (misma de siempre + columna Id Escenario)
        Id Escenario | Id Concepto | Cantidad | Clave Unidad | Descripción |
        Clave Producto/Servicio | Valor Unitario | Objeto Impuesto

    Hoja  'impuestos-conceptos' (misma de siempre + columna Id Escenario)
        Id Escenario | Id Concepto | Tipo Impuesto | Local o Federal |
        Retencion o Traslado | Tipo Factor | Tasa o Cuota | Nombre Impuesto
    ─────────────────────────────────────────────────────────────────────────
    Si una hoja NO tiene la columna 'Id Escenario', todas sus filas se tratan
    como pertenecientes al Escenario 1 (retrocompatibilidad total).

    Devuelve una lista de EscenarioData ordenada por Id Escenario.
    """
    wb = openpyxl.load_workbook(path, data_only=True)

    # ── 1. Hoja 'escenarios' → mapa id → nombre ──────────────────────────────
    escenario_nombres: dict = {}
    if "escenarios" in wb.sheetnames:
        ws_e = wb["escenarios"]
        hdrs_e = [
            str(c.value).strip() if c.value is not None else ""
            for c in next(ws_e.iter_rows(min_row=1, max_row=1))
        ]
        for row in ws_e.iter_rows(min_row=2, values_only=True):
            if not any(row):
                continue
            r = dict(zip(hdrs_e, row))
            eid = int(r.get("Id Escenario") or 0)
            if eid:
                escenario_nombres[eid] = {
                    "nombre":     _clean_str(r.get("Nombre Escenario") or r.get("Nombre") or ""),
                    "rfc_emisor": _clean_str(r.get("RFC Emisor") or ""),
                    "sucursal":   _clean_str(r.get("Sucursal") or ""),
                    "cc":         _clean_str(r.get("CC") or ""),
                }

    # ── 2. Hoja 'conceptos' → agrupar por escenario ──────────────────────────
    ws_c = wb["conceptos"]
    hdrs_c = [
        str(c.value).strip() if c.value is not None else ""
        for c in next(ws_c.iter_rows(min_row=1, max_row=1))
    ]
    has_esc_c = "Id Escenario" in hdrs_c
    conceptos_by_esc: dict = {}

    for row in ws_c.iter_rows(min_row=2, values_only=True):
        if not any(row):
            continue
        r = dict(zip(hdrs_c, row))
        eid = int(r.get("Id Escenario") or 1) if has_esc_c else 1
        desc = _clean_str(r.get("Descripción") or r.get("Descripcion") or "").strip()
        conceptos_by_esc.setdefault(eid, []).append(ConceptoRow(
            cantidad=float(r.get("Cantidad") or 1),
            clave_unidad=_clean_str(r.get("Clave Unidad") or "").upper(),
            descripcion=desc,
            clave_producto=_clean_str(r.get("Clave Producto/Servicio") or "").strip(),
            valor_unitario=float(r.get("Valor Unitario") or 0),
            objeto_impuesto=_normalize_objeto_impuesto(r.get("Objeto Impuesto") or "02"),
            id_concepto=int(r.get("Id Concepto") or 0),
        ))

    # ── 3. Hoja 'impuestos-conceptos' → agrupar por escenario ────────────────
    ws_i = wb["impuestos-conceptos"]
    hdrs_i = [
        str(c.value).strip() if c.value is not None else ""
        for c in next(ws_i.iter_rows(min_row=1, max_row=1))
    ]
    has_esc_i = "Id Escenario" in hdrs_i
    impuestos_by_esc: dict = {}

    for row in ws_i.iter_rows(min_row=2, values_only=True):
        if not any(row):
            continue
        r = dict(zip(hdrs_i, row))
        eid = int(r.get("Id Escenario") or 1) if has_esc_i else 1

        raw_ret = _get_first_present(r, "Retencion o Traslado", "Retención o Traslado")
        raw_imp = _get_first_present(r, "Tipo Impuesto")
        raw_lf  = _get_first_present(r, "Local o Federal")
        raw_tf  = _get_first_present(r, "Tipo Factor")
        raw_tc  = _get_first_present(r, "Tasa o Cuota", default=0)

        if not raw_ret or not raw_imp:
            continue  # fila incompleta — saltar silenciosamente

        try:
            impuestos_by_esc.setdefault(eid, []).append(ImpuestoRow(
                id_concepto=int(_get_first_present(r, "Id Concepto", default=0)),
                retencion_traslado=_norm_retencion(raw_ret),
                clave_impuesto=_clean(raw_imp).upper(),
                nombre_impuesto=_clean(_get_first_present(r, "Nombre Impuesto")),
                local_federal=_norm_local(raw_lf),
                tipo_factor=_norm_factor(raw_tf),
                tasa_cuota=float(raw_tc or 0),
            ))
        except ValueError as exc:
            raise ValueError(
                f"read_escenarios: escenario {eid}, fila {r}: {exc}"
            ) from exc

    wb.close()

    # ── 4. Reunir todos los IDs y construir la lista ordenada ─────────────────
    all_ids = sorted(
        set(list(conceptos_by_esc.keys()) + list(impuestos_by_esc.keys()))
    )
    if not all_ids:
        return []

    result: List[EscenarioData] = []
    for eid in all_ids:
        info = escenario_nombres.get(eid)
        if isinstance(info, dict):
            nombre     = info.get("nombre") or f"Escenario {eid}"
            rfc_emisor = info.get("rfc_emisor", "")
            sucursal   = info.get("sucursal", "")
            cc         = info.get("cc", "")
        else:
            # retrocompatibilidad: antes se almacenaba solo el nombre
            nombre     = info or f"Escenario {eid}"
            rfc_emisor = sucursal = cc = ""
        result.append(EscenarioData(
            id_escenario=eid,
            nombre=nombre,
            conceptos=conceptos_by_esc.get(eid, []),
            impuestos=impuestos_by_esc.get(eid, []),
            rfc_emisor=rfc_emisor,
            sucursal=sucursal,
            cc=cc,
        ))

    return result


def read_pagos(path: str) -> "dict":
    """Lee la hoja 'pagos' del Excel y devuelve un diccionario {id_escenario: PagoRow}.

    Columnas esperadas (fila 1 = encabezados):
        ESCENARIO | FECHA DE PAGO | FORMA DE PAGO | MONEDA DE PAGO | CP1 | CP2

    Si la hoja no existe o está vacía devuelve un diccionario vacío.
    """
    wb = openpyxl.load_workbook(path, data_only=True)
    if "pagos" not in wb.sheetnames:
        wb.close()
        return {}

    ws_p = wb["pagos"]
    header_row = next(ws_p.iter_rows(min_row=1, max_row=1), None)
    if header_row is None:
        wb.close()
        return {}

    headers_p = [
        str(c.value).strip().upper() if c.value is not None else ""
        for c in header_row
    ]

    result: dict = {}
    for row in ws_p.iter_rows(min_row=2, values_only=True):
        if not any(row):
            continue
        r = dict(zip(headers_p, row))

        raw_esc = r.get("ESCENARIO") or r.get("ID ESCENARIO") or r.get("ESC")
        if raw_esc is None:
            continue  # fila sin escenario asignado — saltar
        try:
            eid = int(raw_esc)
        except (TypeError, ValueError):
            continue

        result[eid] = PagoRow(
            id_escenario=eid,
            fecha_pago=_clean_str(
                r.get("FECHA DE PAGO") or r.get("FECHA") or ""
            ),
            forma_pago=_clean_str(
                r.get("FORMA DE PAGO") or r.get("FORMA") or "01"
            ) or "01",
            moneda_pago=(_clean_str(
                r.get("MONEDA DE PAGO") or r.get("MONEDA") or "MXN"
            ) or "MXN").upper(),
            cp1=float(r.get("CP1") or 0),
            cp2=float(r.get("CP2") or 0),
        )

    wb.close()
    return result


def get_pago_by_escenario(pagos: dict, id_escenario: int) -> "PagoRow":
    """Devuelve el PagoRow del escenario solicitado.

    Lanza KeyError con mensaje claro si no existe entrada para ese escenario,
    para que el test falle con un mensaje descriptivo.
    """
    if id_escenario not in pagos:
        raise KeyError(
            f"No hay datos de pagos para el escenario {id_escenario}. "
            f"Escenarios disponibles en la hoja 'pagos': {sorted(pagos.keys())}"
        )
    return pagos[id_escenario]


# ── Writer ────────────────────────────────────────────────────────────────────

class ResultsWriter:
    """Creates / appends rows to the results Excel file."""

    def __init__(self, reports_dir: str):
        os.makedirs(reports_dir, exist_ok=True)
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        self.path = os.path.join(reports_dir, f"resultados_{ts}.xlsx")
        self._wb = openpyxl.Workbook()
        self._ws = self._wb.active
        self._ws.title = "Resultados"
        self._write_headers()

    def _write_headers(self):
        for col, header in enumerate(_HEADERS, start=1):
            cell = self._ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF", size=11)
            cell.fill = _HEADER_FILL
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = _THIN_BORDER
        self._ws.row_dimensions[1].height = 30

    def add_row(self, result: ResultRow):
        row_idx = self._ws.max_row + 1
        values = [getattr(result, f.name) for f in fields(result)]
        is_pass = str(result.resultado_obtenido).upper().startswith("PASS")
        fill = _PASS_FILL if is_pass else _FAIL_FILL

        for col, value in enumerate(values, start=1):
            cell = self._ws.cell(row=row_idx, column=col, value=value)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.border = _THIN_BORDER
            if col == len(_HEADERS):
                cell.fill = fill

    def save(self):
        for col_cells in self._ws.columns:
            max_len = max((len(str(c.value or "")) for c in col_cells), default=10)
            self._ws.column_dimensions[get_column_letter(col_cells[0].column)].width = min(max_len + 4, 60)
        self._wb.save(self.path)
        return self.path