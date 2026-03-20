"""
Script to generate the test Excel files.
Run once: python create_test_data.py

Creates:
  - data/conceptos-escenario-ppd-1.xlsx (conceptos + impuestos sheets)
"""
import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

BASE = os.path.dirname(os.path.abspath(__file__))
DATA_DIR = os.path.join(BASE, "data")
os.makedirs(DATA_DIR, exist_ok=True)

CONCEPTOS_PATH = os.path.join(DATA_DIR, "conceptos-escenario-ppd-1.xlsx")


def _header_row(ws, headers):
    fill = PatternFill("solid", fgColor="2D5F9E")
    font = Font(bold=True, color="FFFFFF")
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.fill = fill
        c.font = font
        c.alignment = Alignment(horizontal="center")
    ws.row_dimensions[1].height = 20


def create_conceptos_excel():
    wb = openpyxl.Workbook()

    # ── Sheet 1: conceptos ────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "conceptos"
    headers1 = ["Cantidad", "Clave Unidad", "Descripción",
                 "Clave Producto/Servicio", "Valor Unitario", "Objeto Impuesto", "Id Concepto"]
    _header_row(ws1, headers1)

    data1 = [
        (1,   "E48", "PRUEBAS",  "01010101", 1000.00, "02", 1),
        (2,   "E48", "PRUEBAS",  "01010101", 1000.00, "02", 2),
    ]
    for row in data1:
        ws1.append(row)

    # Auto-width
    for col in ws1.columns:
        ws1.column_dimensions[col[0].column_letter].width = max(
            len(str(c.value or "")) for c in col) + 4

    # ── Sheet 2: impuestos-conceptos ──────────────────────────────────────
    ws2 = wb.create_sheet("impuestos-conceptos")
    headers2 = ["Id Concepto", "Tipo Impuesto", "Clave Impuesto",
                 "Nombre Impuesto", "Tipo Factor", "Tasa o Cuota"]
    _header_row(ws2, headers2)

    data2 = [
        (1, "Traslado", "IVA", "IVA FEDERAL TRASLADO TASA 0.16", "Tasa", 0.16),
        (2, "Traslado", "IVA", "IVA FEDERAL TRASLADO TASA 0.16", "Tasa", 0.16),
    ]
    for row in data2:
        ws2.append(row)

    for col in ws2.columns:
        ws2.column_dimensions[col[0].column_letter].width = max(
            len(str(c.value or "")) for c in col) + 4

    wb.save(CONCEPTOS_PATH)
    print(f"✅ Created: {CONCEPTOS_PATH}")


if __name__ == "__main__":
    create_conceptos_excel()
    print("\nDone. Update config/config.ini → [data] conceptos_excel if needed.")
